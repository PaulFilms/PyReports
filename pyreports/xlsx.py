'''
Toolkit with simplified functions and methods for create .xlsx spreadsheets
'''
__update__ = '2024.11.28'

import os
import re
import locale
from typing import List, Union, Type, Any
from enum import Enum

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, borders, PatternFill, Protection
from openpyxl.worksheet import pagebreak
from openpyxl.utils import get_column_letter, quote_sheetname, absolute_coordinate
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.drawing.image import Image

# from PIL import Image as PILImage


## TOOLS
## _________________________________________________________________________________________________________________

class alignments_vertical(Enum):
    '''
    vertical alignments types
    '''
    top = "top"
    center = "center"
    bottom = "bottom"
    justify = "justify"
    distributed = "distributed"

class alignments_horizontal(Enum):
    '''
    horizontal alignments types
    '''
    general = "general"
    left = "left"
    center = "center"
    right = "right"
    fill = "fill"
    justify = "justify"
    center_continuous = "centerContinuous"
    distributed = "distributed"

class alignments(Enum):
    '''
    Normalized alignments
    '''
    main = Alignment(horizontal=alignments_horizontal.left.value, vertical=alignments_vertical.center.value)
    top = Alignment(horizontal=alignments_horizontal.left.value, vertical=alignments_vertical.top.value)

class fonts(Enum):
    '''
    Normalized Fonts
    '''
    title = Font(name='Calibri', size=12, bold=True)
    header = Font(name='Calibri', size=10, bold=True)
    main = Font(name='Calibri', size=10, bold=False)
    caption = Font(name='Calibri', size=8, bold=False)

class pattern_fills(Enum):
    RED = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    GREY = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
    GREEN = PatternFill(start_color='93DC5C', end_color='93DC5C', fill_type='solid')
    YELLOW = PatternFill(start_color='E5DE00', end_color='E5DE00', fill_type='solid')
    BLUE = PatternFill(start_color='1A43BF', end_color='1A43BF', fill_type='solid')


## FUNCTIONS
## _________________________________________________________________________________________________________________

def get_os_decimal() -> str:
    '''
    Get the decimal separator used by the Operating System
    '''
    return str(locale.localeconv()['decimal_point'])

def get_cell(column: int, row: int = None) -> str:
    '''
    Get Column or cell reference

    `Args:`
        - column: int
        - row: int (optional)
    
    `Returns:`
        - column = 1 -> 'A'
        - column = 6 -> 'F'
        - column = 6, row = 7 -> 'F7'
    '''
    string = str()
    while column > 0:
        column, remainder = divmod(column - 1, 26)
        string = chr(65 + remainder) + string
    if row:
        string += str(row)
    return string

def get_formula(formula: str, row: int, columns_enum: Type[Enum]) -> str:
    '''
    Get a formula by replacing the variables into "<<>>" separators with the corresponding reference and row

    `Args:`
        - formula: str
        - row: int
        - columns_enum: Type[Any]

    ***Note:*** `columns_enum` must be defined as an Enum class, like this:
    ```python
    class my_columns(Enum):
        col1 = 1
        col2 = 6
    ```
    
    `Markers:`
    ```
    <<ROW>> ** Fixed marker reserved to row indication
    <<COLUMN_NAME>>
    <<COLUMN_NAME+3>> ** With column (+)offset
    <<COLUMN_NAME-4>> ** With column (-)offset
    ```
    
    `Examples:`
    ```
    '=IF(<<col2>><<ROW>><0,"Pass","Fail")'
    '=COUNT(<<col1>><<ROW>>:<<col1+9>><<ROW>>)'
    ```
    '''
    if '<<ROW>>' in formula:
        formula = formula.replace('<<ROW>>', str(row))
    for col in columns_enum:
        placeholder = f"<<{col.name}>>"
        column_letter = get_cell(col.value)
        formula = formula.replace(placeholder, column_letter)
        ## OFFSET
        Match = re.search(rf"<<{col.name}([+-]\d+)?>>", formula)
        if Match:
            offset = int(Match.group(1))
            column_letter = get_cell(col.value+offset)
            formula = formula.replace(Match.group(0), column_letter)
    return formula


## XLSX REPORT
## _________________________________________________________________________________________________________________

class XLSREPORT:
    '''
    Sub-Module to make and edit .xlsx Reports

    `Args:`
        - `path` (str): Complete or relative path of report file
        - `worksheet_name` (str): Name of current DataSheet
    '''

    def __init__(self, path: str, worksheet_name: str = None) -> None: # worksheet_name: str = "Data"
        self.filePath = path
        extension = os.path.splitext(path)[1]
        if extension == str() or extension == None:
            self.filePath += ".xlsx"

        ## NEW WORKBOOK
        if not os.path.exists(self.filePath):
            self.wb = Workbook(self.filePath)
            self.wb.create_sheet(worksheet_name)
            if worksheet_name:
                self.ws = self.wb[worksheet_name]
            else:
                self.ws = self.wb['Sheet1']
            self.ws.sheet_format.defaultRowHeight = 15
            self.wb.save(self.filePath)
            self.wb.close()

        ## LOAD WORKBOOK
        self.wb = load_workbook(self.filePath, read_only=False) ## Force load_workbook to open like WriteOnly

        ## WORKSHEET
        if not worksheet_name:
            self.ws = self.wb[self.wb.sheetnames[0]]
        else:
            if worksheet_name in self.wb.sheetnames:
                self.ws = self.wb[worksheet_name]
            else:
                self.wb.create_sheet(worksheet_name)
                self.ws = self.wb[worksheet_name]
                self.wb.save(self.filePath)

        ## INIT
        self.row: int = 1

    def save(self) -> None:
        self.wb.save(self.filePath)

    def close(self) -> None:
        self.wb.save(self.filePath)
        self.wb.close()

    def get_properties(self) -> any:
        return self.wb.properties

    def sheet_list(self) -> List[str]:
        return self.wb.sheetnames 

    def sheet_select(self, sheet_name: str) -> None:
        self.ws = self.wb[sheet_name]

    def sheet_new(self, sheet_name: str) -> None:
        '''
        Create and select a new excel sheet
        '''
        if not sheet_name in self.wb.sheetnames:
            self.wb.create_sheet(sheet_name)
        self.ws = self.wb[sheet_name]
        self.ws.sheet_format.defaultRowHeight = 15
        # self.ws.row_dimensions[1:].height = 15

    def row_inc(self, number: int = 1) -> None:
        '''
        Add an increment in row count
        '''
        self.row += int(number)

    def row_height(self, row: int, height: float = 10) -> None:
        '''
        Set height of a row
        '''
        self.ws.row_dimensions[row].height = height

    def col_width(self, column: int, width: float = 20) -> None:
        '''
        Set width of a column
        '''
        self.ws.column_dimensions[get_column_letter(column)].width = width

    def col_autofit(self) -> None:
        '''
        Auto-Adjust the Column Width 
        '''
        for column_cells in self.ws.columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = (get_column_letter(column_cells[0].column))
                if new_column_length > 0:
                    self.ws.column_dimensions[new_column_letter].width = new_column_length*1.23

    def col_filters(self) -> None:
        '''
        Set filters in current WorkSheet from A1 to maximun column and maximun row
        '''
        fullRange = f"A1:{get_column_letter(self.ws.max_column)}{self.ws.max_row}"
        self.ws.auto_filter.ref = fullRange
    
    def cell_protect(self, row: int, column: int) -> None:
        self.ws.cell(row, column).protection = Protection(locked=True)


    ## READ/WRITE FUNCTIONS
    ## _________________________________________________________________________________________________________________

    def rd(self, row: int, column: int) -> any:
        '''
        Returns value of selected row and column from current sheet
        '''
        return self.ws.cell(row, column).value

    def wr(self, row: int, column: int, value: any = None, font: Font = fonts.main.value, alignment: Alignment = alignments.main.value) -> None:
        '''
        Type the selected cell in specific formatting
        - `size:` Font Size (10)
        - `bold:` Font Bold (False)
        - `font_name:` Font Name ("Arial")
        '''
        try:
            self.ws.cell(row, column).value = value
            self.ws.cell(row, column).alignment = alignment
            self.ws.cell(row, column).font = font
        except Exception as e:
            print("ERROR wr:")
            print(e)
            self.ws.cell(row, column).value = "ERROR"
        # self.row_height(row, 15)

    def wr_title(self, row: int, column: int, value: str):
        '''
        Write selected cell with Title format
        '''
        self.wr(row, column, value, font=fonts.title.value, alignment=alignments.main.value)
        self.row_height(row, 30)

    def wr_header(self, row: int, column: int, value: str, wrap_text: bool = False) -> None:
        '''
        Write selected cell with Header format
        '''
        self.wr(row, column, value, font=fonts.header.value, alignment=alignments.main.value)
        self.ws.cell(row, column).alignment = Alignment(
            horizontal=alignments_horizontal.left.value,
            vertical=alignments_vertical.center.value,
            wrap_text=wrap_text
        )
        # self.row_height(row, 30)

    def wr_headers(self, row: int, column_init: int, headers: List[str], wrap_text: bool = False) -> None:
        '''
        Write selected cell with Header format
        '''
        col = column_init
        for header in headers:
            self.wr_header(row=row, column=col, value=header, wrap_text=wrap_text)
            col+=1
        self.row_height(row, 30)

    def wr_sci_number(self, row: int, column: int, value: int | float) -> None:
        '''
        Edit selected cell like sci number format (0.0E+0)
        '''
        self.wr(row, column, value)
        self.ws.cell(row, column).number_format = '0.0E+0'
        # self.row_height(row, 15)

    def wr_image(self, row: int, column: int, img_path: str, scale: float = 100.0) -> None:
        '''
        Insert an image (*.jpg, *.png) into the selected cell 
        '''
        image = Image(img_path)
        image.height = image.height * scale / 100
        image.width = image.width * scale / 100
        self.ws.add_image(image, get_cell(column=column, row=row))


    ## UNDER TEST
    ## _________________________________________________________________________________________________________________

    def warp(self, row: int, column: int) -> None:
        '''
        '''
        self.ws.cell(row, column).alignment = Alignment(wrap_text=True)

    def set_range_name(self, row: int, column: int, name: str) -> None:
        '''
        BUG: Under Test
        '''
        cell = f"{get_column_letter(column)}{row}"
        ref =  f"{quote_sheetname(self.ws.title)}!{absolute_coordinate(cell)}"
        defn = DefinedName(name, attr_text=ref)
        self.wb.defined_names[name] = defn

    def low_border(self, row: int, col_ini: int = 1, col_fin: int = 300) -> None:
        '''
        BUG: INCOMPLETE
        Hay que saber bien el diseño y todas las funciones de borders

        https://openpyxl.readthedocs.io/en/stable/styles.html?highlight=border_style
        '''
        # Style = "thick" (Grueso)
        border0 = borders.Side(style = None, color = None, border_style = None)
        borderLow = borders.Side(
            style = "medium", 
            color="000000", 
            # border_style = "double"
            )
        thin = borders.Border(left = border0, right = border0, bottom = borderLow, top = border0)
        for col in range(col_ini, col_fin): 
            self.ws.cell(row=row, column=col).border = thin

    def sheet_print_area(self, column_fin: int | str) -> None:
        '''
        Ajusta la zona de impresion
        INCOMPLETE
        '''
        self.ws.page_setup.fitToPage = 1
        self.ws.page_setup.fitToHeight = False
        if isinstance(column_fin, int):
            COL_STR = get_column_letter(column_fin)
        if isinstance(column_fin, str):
            COL_STR = column_fin
        self.ws.print_area = "A:" + COL_STR

    def sheet_head(self, row_fin: int) -> None:
        '''
        Define la cabecera superior
        '''
        self.ws.print_title_rows = "1:" + str(row_fin)
        self.ws.page_margins.top = 0.4
        self.ws.page_margins.botom = 0.4
        # self.ws.page_margins.header = 0.7
        self.ws.page_margins.header = 0.4
        self.ws.page_margins.footer = 0.4

    def page_break(self, row: int = 1) -> None:
        '''
        Insert a page break in selected row
        '''
        page_break = pagebreak.Break(id=row-1)
        break_list = self.ws.row_breaks
        break_list.append(page_break)
        self.ws.row_breaks = break_list



## PANDAS
## _________________________________________________________________________________________________________________

import pandas as pd

def DF_REPORT(dataFrame: pd.DataFrame, path: str) -> None:
    '''
    Create excel report from selected Pandas DataFrame
    '''
    report = XLSREPORT(path)
    ## HEADERS
    headers: list = dataFrame.columns.values.tolist()
    report.wr_headers(1, 1, headers)
    report.col_filters()
    report.low_border(report.row, col_fin=len(headers)+1)
    report.row_inc()
    ## DATA
    for row in range(len(dataFrame.index)):
        row_data = list(dataFrame.iloc[row].values)
        for value in row_data:
            report.wr(report.row, row_data.index(value)+1, value)
        report.row_inc()
    report.col_autofit()
    ## FIN
    report.save()
    report.close()